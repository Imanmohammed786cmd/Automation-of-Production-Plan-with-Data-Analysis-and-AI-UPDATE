from __future__ import annotations

import json
import math
import re
from cgi import FieldStorage
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path
from urllib.parse import unquote

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parent
EXPORT_DIR = ROOT / "exports"
DOWNLOADS_DIR = Path.home() / "Downloads"
BOM_CACHE: dict | None = None


def load_bom_data() -> dict:
    global BOM_CACHE
    if BOM_CACHE is not None:
        return BOM_CACHE

    text = (ROOT / "data.js").read_text(encoding="utf-8").strip()
    prefix = "window.BOM_DATA = "
    if text.startswith(prefix):
        text = text[len(prefix) :]
    if text.endswith(";"):
        text = text[:-1]
    data = json.loads(text)
    by_item: dict[str, list[dict]] = {}
    for row in data["records"]:
        by_item.setdefault(str(row.get("i", "")), []).append(row)
    BOM_CACHE = {"raw": data, "by_item": by_item}
    return BOM_CACHE


def add_days(date_value: str, days: int) -> str:
    from datetime import date, timedelta

    try:
        year, month, day = [int(part) for part in date_value.split("-")]
        start = date(year, month, day)
        offset = max(0, int(days or 0))
        max_offset = (date.max - start).days
        if offset > max_offset:
            return f">{date.max.isoformat()}"
        return (start + timedelta(days=offset)).isoformat()
    except Exception:
        return date_value


def machine_sort_value(machine: str) -> int:
    order = [25, 45, 50, 80, 120, 160]
    match = re.search(r"(?:^|\D)(25|45|50|80|120|160)(?:\D|$)", str(machine or "").upper())
    if not match:
        return 1000
    return order.index(int(match.group(1)))


def machine_order_label(machine: str) -> str:
    match = re.search(r"(?:^|\D)(25|45|50|80|120|160)(?:\D|$)", str(machine or "").upper())
    return f"{match.group(1)}T" if match else "Other"


def build_server_plan_rows(payload: dict) -> list[dict]:
    bom = load_bom_data()["by_item"]
    inputs = payload.get("inputs", [])
    production_date = payload.get("productionDate") or "2026-05-10"
    shifts_per_day = max(1, int(payload.get("shiftsPerDay") or 3))

    items = []
    missing = []
    for entry in inputs:
        code = str(entry.get("code", "")).strip()
        qty = float(entry.get("qty") or 0)
        if not code or qty <= 0:
            continue
        bom_rows = bom.get(code)
        if not bom_rows:
            missing.append({"code": code, "qty": qty})
            continue
        first = bom_rows[0]
        hour_qty = float(first.get("hq") or 0)
        cycle_time = float(first.get("ct") or 0)
        hours = qty / hour_qty if hour_qty > 0 else qty * cycle_time
        shifts = math.ceil(hours / 8)
        days = math.ceil(shifts / shifts_per_day)
        items.append(
            {
                "code": code,
                "qty": qty,
                "component": first.get("c", ""),
                "machine": first.get("ma", "") or "Unassigned",
                "hours": hours,
                "days": days,
                "bomRows": bom_rows,
            }
        )

    items.sort(key=lambda item: (machine_sort_value(item["machine"]), item["machine"], -item["hours"]))
    machine_finish_day: dict[str, int] = {}
    rows: list[dict] = []
    for item in items:
        machine = item["machine"]
        offset = machine_finish_day.get(machine, 0)
        start_date = add_days(production_date, offset)
        machine_finish_day[machine] = offset + item["days"]
        for bom_row in item["bomRows"]:
            factor = float(bom_row.get("w") or 0)
            cycle_time = round(float(bom_row.get("ct") or 0) * 60, 6)
            rows.append(
                {
                    "Date": start_date,
                    "Machine": machine,
                    "Item Code": item["code"],
                    "Component": item["component"],
                    "Plan Qty": int(item["qty"]) if item["qty"].is_integer() else item["qty"],
                    "BOM Cycle Time": cycle_time,
                    "Material": bom_row.get("m", ""),
                    "Material Required": round(item["qty"] * factor, 6),
                    "Production Hrs": round(item["hours"], 4),
                }
            )

    for item in missing:
        rows.append(
            {
                "Date": production_date,
                "Machine": "",
                "Item Code": item["code"],
                "Component": "Item code not found in BOM Reference",
                "Plan Qty": item["qty"],
                "BOM Cycle Time": "",
                "Material": "",
                "Material Required": "",
                "Production Hrs": "",
            }
        )

    return rows


def build_server_machine_plan_rows(payload: dict) -> list[dict]:
    bom_data = load_bom_data()
    bom = bom_data["by_item"]
    item_names = {str(item.get("code", "")): item.get("component", "") for item in bom_data["raw"].get("items", [])}
    inputs = payload.get("inputs", [])
    production_date = payload.get("productionDate") or "2026-05-10"
    shifts_per_day = max(1, int(payload.get("shiftsPerDay") or 3))
    generated_at = payload.get("generatedAt") or ""

    items = []
    missing = []
    for entry in inputs:
        code = str(entry.get("code", "")).strip()
        qty = float(entry.get("qty") or 0)
        if not code or qty <= 0:
            continue
        bom_rows = bom.get(code)
        if not bom_rows:
            missing.append({"code": code, "qty": qty})
            continue
        first = bom_rows[0]
        hour_qty = float(first.get("hq") or 0)
        cycle_time = float(first.get("ct") or 0)
        hours = qty / hour_qty if hour_qty > 0 else qty * cycle_time
        shifts = math.ceil(hours / 8)
        days = math.ceil(shifts / shifts_per_day)
        machine = first.get("ma", "") or "Unassigned"
        items.append(
            {
                "code": code,
                "qty": qty,
                "component": first.get("c", ""),
                "machine": machine,
                "process": first.get("p", ""),
                "hours": hours,
                "shifts": shifts,
                "days": days,
            }
        )

    items.sort(key=lambda item: (machine_sort_value(item["machine"]), item["machine"], -item["hours"]))
    machine_finish_day: dict[str, int] = {}
    rows: list[dict] = []
    for item in items:
        machine = item["machine"]
        offset = machine_finish_day.get(machine, 0)
        start_date = add_days(production_date, offset)
        end_date = add_days(start_date, max(0, item["days"] - 1))
        machine_finish_day[machine] = offset + item["days"]
        qty = int(item["qty"]) if item["qty"].is_integer() else item["qty"]
        rows.append(
            {
                "Machine Order": machine_order_label(machine),
                "Machine": machine,
                "Process": item["process"],
                "Item Code": item["code"],
                "Component": item["component"],
                "Production Qty": qty,
                "Start Date": start_date,
                "End Date": end_date,
                "Required Hrs": round(item["hours"], 4),
                "Shifts": item["shifts"],
                "Shifts per Day": shifts_per_day,
                "Days to Complete": item["days"],
                "Generated At": generated_at,
                "Status": "OK",
            }
        )

    for item in missing:
        rows.append(
            {
                "Machine Order": "",
                "Machine": "",
                "Process": "",
                "Item Code": item["code"],
                "Component": item_names.get(item["code"], ""),
                "Production Qty": int(item["qty"]) if item["qty"].is_integer() else item["qty"],
                "Start Date": production_date,
                "End Date": "",
                "Required Hrs": "",
                "Shifts": "",
                "Shifts per Day": shifts_per_day,
                "Days to Complete": "",
                "Generated At": generated_at,
                "Status": "Item code not found in BOM Reference",
            }
        )

    return rows


class PlannerHandler(SimpleHTTPRequestHandler):
    def end_headers(self) -> None:
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.send_header("Cache-Control", "no-store, no-cache, must-revalidate")
        super().end_headers()

    def do_OPTIONS(self) -> None:
        self.send_response(204)
        self.end_headers()

    def do_GET(self) -> None:
        if self.path.startswith("/exports/"):
            self.serve_export_file()
            return
        super().do_GET()

    def serve_export_file(self) -> None:
        filename = Path(unquote(self.path.split("/exports/", 1)[1])).name
        path = EXPORT_DIR / filename
        if not path.exists() or not path.is_file():
            self.send_error(404, "Export file not found")
            return

        if filename.endswith(".xlsx"):
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        elif filename.endswith(".csv"):
            content_type = "text/csv; charset=utf-8"
        else:
            content_type = "application/octet-stream"

        data = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.end_headers()
        self.wfile.write(data)

    def do_POST(self) -> None:
        if self.path == "/export-plan-xlsx":
            self.export_plan_xlsx()
            return

        if self.path == "/export-machine-plan-xlsx":
            self.export_machine_plan_xlsx()
            return

        if self.path == "/save-plan-xlsx":
            self.save_plan_xlsx()
            return

        if self.path == "/save-export":
            self.save_export()
            return

        if self.path != "/parse-upload":
            self.send_json({"error": "Not found"}, status=404)
            return

        try:
            form = FieldStorage(
                fp=self.rfile,
                headers=self.headers,
                environ={
                    "REQUEST_METHOD": "POST",
                    "CONTENT_TYPE": self.headers.get("Content-Type", ""),
                    "CONTENT_LENGTH": self.headers.get("Content-Length", "0"),
                },
            )
            upload = form["file"] if "file" in form else None
            if upload is None or not getattr(upload, "filename", ""):
                self.send_json({"error": "No file uploaded."}, status=400)
                return

            filename = upload.filename.lower()
            content = upload.file.read()
            if not filename.endswith(".xlsx"):
                self.send_json({"error": "Only .xlsx Excel files are supported. Save .xls as .xlsx and upload again."}, status=400)
                return

            rows = self.read_xlsx(content)
            self.send_json({"rows": rows, "rowCount": len(rows)})
        except Exception as exc:
            self.send_json({"error": f"Unable to parse Excel file: {exc}"}, status=400)

    def save_export(self) -> None:
        try:
            length = int(self.headers.get("Content-Length", "0"))
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
            filename = re.sub(r"[^A-Za-z0-9_.-]+", "_", payload.get("filename", "export.csv")).strip("._")
            if not filename.endswith(".csv"):
                filename += ".csv"
            content = payload.get("content", "")
            if not content:
                self.send_json({"error": "No export content received."}, status=400)
                return

            EXPORT_DIR.mkdir(exist_ok=True)
            path = EXPORT_DIR / filename
            path.write_text(content, encoding="utf-8")
            saved_path = path
            try:
                downloads_path = DOWNLOADS_DIR / filename
                downloads_path.write_text(content, encoding="utf-8")
                saved_path = downloads_path
            except Exception:
                pass
            self.send_json({"url": f"/exports/{filename}", "filename": filename, "path": str(saved_path)})
        except Exception as exc:
            self.send_json({"error": f"Unable to save export: {exc}"}, status=400)

    def save_plan_xlsx(self) -> None:
        try:
            length = int(self.headers.get("Content-Length", "0"))
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
            filename = re.sub(r"[^A-Za-z0-9_.-]+", "_", payload.get("filename", "machine_wise_plan.xlsx")).strip("._")
            if not filename.endswith(".xlsx"):
                filename += ".xlsx"
            rows = payload.get("rows", [])
            if not rows:
                self.send_json({"error": "No plan rows received."}, status=400)
                return
            self.write_plan_workbook(filename, rows)
        except Exception as exc:
            self.send_json({"error": f"Unable to save Excel plan: {exc}"}, status=400)

    def export_plan_xlsx(self) -> None:
        try:
            length = int(self.headers.get("Content-Length", "0"))
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
            rows = build_server_plan_rows(payload)
            if not rows:
                self.send_json({"error": "No production rows available to export."}, status=400)
                return
            self.write_plan_workbook("machine_wise_production_plan.xlsx", rows)
        except Exception as exc:
            self.send_json({"error": f"Unable to generate Excel plan: {exc}"}, status=400)

    def export_machine_plan_xlsx(self) -> None:
        try:
            length = int(self.headers.get("Content-Length", "0"))
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
            rows = build_server_machine_plan_rows(payload)
            if not rows:
                self.send_json({"error": "No production rows available to export."}, status=400)
                return
            self.write_plan_workbook("machine_wise_plan_25_to_160.xlsx", rows)
        except Exception as exc:
            self.send_json({"error": f"Unable to generate machine plan: {exc}"}, status=400)

    def write_plan_workbook(self, filename: str, rows: list[dict]) -> None:
        headers = list(rows[0].keys())
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Machine Wise Plan"
        sheet.append(headers)

        for row in rows:
            sheet.append([row.get(header, "") for header in headers])

        header_fill = PatternFill("solid", fgColor="0F6B5E")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        widths = {
            "Date": 14,
            "Machine": 24,
            "Item Code": 18,
            "Component": 44,
            "Plan Qty": 14,
            "BOM Cycle Time": 16,
            "Material": 36,
            "Material Required": 18,
            "Production Hrs": 18,
        }
        for idx, header in enumerate(headers, start=1):
            letter = sheet.cell(row=1, column=idx).column_letter
            sheet.column_dimensions[letter].width = widths.get(header, 18)

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        EXPORT_DIR.mkdir(exist_ok=True)
        path = EXPORT_DIR / filename
        workbook.save(path)
        saved_path = path
        try:
            downloads_path = DOWNLOADS_DIR / filename
            workbook.save(downloads_path)
            saved_path = downloads_path
        except Exception:
            pass
        self.send_json({"url": f"/exports/{filename}", "filename": filename, "path": str(saved_path)})

    def read_xlsx(self, content: bytes) -> list[list[str]]:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
        sheet = workbook.active
        rows: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            values = ["" if value is None else str(value).strip() for value in row]
            while values and values[-1] == "":
                values.pop()
            if any(values):
                rows.append(values)
        return rows

    def send_json(self, payload: dict, status: int = 200) -> None:
        body = json.dumps(payload).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)


def main() -> None:
    server = ThreadingHTTPServer(("127.0.0.1", 4175), lambda *args, **kwargs: PlannerHandler(*args, directory=str(ROOT), **kwargs))
    print("Production planner running at http://127.0.0.1:4175/index.html")
    server.serve_forever()


if __name__ == "__main__":
    main()
