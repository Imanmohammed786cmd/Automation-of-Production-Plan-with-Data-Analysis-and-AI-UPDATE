from __future__ import annotations

import json
import math
import re
from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parent
INPUT_FILE = Path("/Users/imanmohammed/Downloads/May Plan.xlsx")
EXPORT_DIR = ROOT / "exports"
DOWNLOADS_DIR = Path.home() / "Downloads"
OUTPUT_NAME = "may_machine_wise_plan.xlsx"
PRODUCTION_DATE = "2026-05-13"
SHIFTS_PER_DAY = 3


def norm(value: object) -> str:
    text = str(value or "").replace("\xa0", " ")
    return re.sub(r"\s+", " ", text).strip().upper()


def load_bom() -> dict:
    text = (ROOT / "data.js").read_text(encoding="utf-8").strip()
    prefix = "window.BOM_DATA = "
    if text.startswith(prefix):
        text = text[len(prefix) :]
    if text.endswith(";"):
        text = text[:-1]
    data = json.loads(text)
    by_item: dict[str, list[dict]] = defaultdict(list)
    first_by_item: dict[str, dict] = {}
    by_component: dict[str, list[str]] = defaultdict(list)

    for row in data["records"]:
        code = str(row.get("i", "")).strip()
        by_item[code].append(row)
        first_by_item.setdefault(code, row)

    for item in data.get("items", []):
        component = norm(item.get("component", ""))
        code = str(item.get("code", "")).strip()
        if component and code:
            by_component[component].append(code)

    return {"raw": data, "by_item": by_item, "first_by_item": first_by_item, "by_component": by_component}


def add_days(date_value: str, days: int) -> str:
    start = date.fromisoformat(date_value)
    return (start + timedelta(days=max(0, int(days or 0)))).isoformat()


def machine_sort_value(machine: str) -> int:
    order = [25, 45, 50, 80, 120, 160]
    match = re.search(r"(?:^|\D)(25|45|50|80|120|160)(?:\D|$)", str(machine or "").upper())
    if not match:
        return 1000
    return order.index(int(match.group(1)))


def machine_order_label(machine: str) -> str:
    match = re.search(r"(?:^|\D)(25|45|50|80|120|160)(?:\D|$)", str(machine or "").upper())
    return f"{match.group(1)}T" if match else "Other"


def choose_code(codes: list[str], first_by_item: dict[str, dict]) -> str:
    def score(code: str) -> tuple:
        row = first_by_item.get(code, {})
        return (
            machine_sort_value(row.get("ma", "")),
            0 if not code.startswith("50-") else 1,
            code,
        )

    return sorted(codes, key=score)[0]


def read_may_plan() -> list[dict]:
    workbook = load_workbook(INPUT_FILE, read_only=True, data_only=True)
    sheet = workbook.active
    header_row = None
    header_values = []
    for row_no, values in enumerate(sheet.iter_rows(values_only=True), start=1):
        normalized = [norm(value) for value in values]
        if "ITEM DESCRIPTION" in normalized and "BALANCE" in normalized:
            header_row = row_no
            header_values = normalized
            break
    if header_row is None:
        raise ValueError("Could not find Item Description and Balance headers in May Plan.xlsx")

    description_index = header_values.index("ITEM DESCRIPTION")
    balance_index = header_values.index("BALANCE")
    code_index = description_index - 1 if description_index > 0 else None

    rows = []
    for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        description = values[description_index] if description_index < len(values) else ""
        balance = values[balance_index] if balance_index < len(values) else 0
        code = values[code_index] if code_index is not None and code_index < len(values) else ""
        if not description:
            continue
        try:
            qty = float(balance or 0)
        except Exception:
            qty = 0
        if qty > 0:
            rows.append({"code": str(code or "").strip(), "description": str(description).strip(), "qty": qty})
    return rows


def build_rows() -> tuple[list[dict], list[dict], list[dict]]:
    bom = load_bom()
    by_item = bom["by_item"]
    by_component = bom["by_component"]
    first_by_item = bom["first_by_item"]

    input_rows = read_may_plan()
    qty_by_code: dict[str, float] = defaultdict(float)
    match_rows = []

    for line_no, row in enumerate(input_rows, start=2):
        input_code = row.get("code", "").strip()
        description = row["description"]
        qty = row["qty"]
        codes = by_component.get(norm(description), [])
        if input_code and input_code in by_item:
            chosen = input_code
            if input_code not in codes:
                codes = [input_code, *codes]
        else:
            chosen = choose_code(codes, first_by_item) if codes else ""
        if codes:
            qty_by_code[chosen] += qty
            first = first_by_item.get(chosen, {})
            if input_code and input_code in by_item:
                status = "Input item code used"
            else:
                status = "OK" if len(codes) == 1 else "Multiple matches - selected best machine code"
            match_rows.append(
                {
                    "Source Row": line_no,
                    "Item Description": description,
                    "Balance": int(qty) if qty.is_integer() else qty,
                    "Input Item Code": input_code,
                    "Selected Item Code": chosen,
                    "Selected Machine": first.get("ma", ""),
                    "Selected Process": first.get("p", ""),
                    "Alternate Item Codes": ", ".join(codes),
                    "Status": status,
                }
            )
        else:
            match_rows.append(
                {
                    "Source Row": line_no,
                    "Item Description": description,
                    "Balance": int(qty) if qty.is_integer() else qty,
                    "Input Item Code": input_code,
                    "Selected Item Code": "",
                    "Selected Machine": "",
                    "Selected Process": "",
                    "Alternate Item Codes": "",
                    "Status": "Description not found in BOM Reference",
                }
            )

    items = []
    for code, qty in qty_by_code.items():
        bom_rows = by_item.get(code, [])
        if not bom_rows:
            continue
        first = bom_rows[0]
        hour_qty = float(first.get("hq") or 0)
        cycle_time = float(first.get("ct") or 0)
        hours = qty / hour_qty if hour_qty > 0 else qty * cycle_time
        shifts = math.ceil(hours / 8)
        days = math.ceil(shifts / SHIFTS_PER_DAY)
        items.append(
            {
                "code": code,
                "qty": qty,
                "component": first.get("c", ""),
                "machine": first.get("ma", "") or "Unassigned",
                "process": first.get("p", ""),
                "hours": hours,
                "shifts": shifts,
                "days": days,
            }
        )

    items.sort(key=lambda item: (machine_sort_value(item["machine"]), item["machine"], -item["hours"]))
    machine_finish_day: dict[str, int] = {}
    plan_rows = []
    for item in items:
        machine = item["machine"]
        offset = machine_finish_day.get(machine, 0)
        start_date = add_days(PRODUCTION_DATE, offset)
        end_date = add_days(start_date, max(0, item["days"] - 1))
        machine_finish_day[machine] = offset + item["days"]
        qty = int(item["qty"]) if item["qty"].is_integer() else item["qty"]
        plan_rows.append(
            {
                "Machine Order": machine_order_label(machine),
                "Machine": machine,
                "Process": item["process"],
                "Item Code": item["code"],
                "Component": item["component"],
                "Production Qty": qty,
                "Start Date": start_date,
                "End Date": end_date,
                "Required Hrs": round(item["hours"], 4),
                "Shifts": item["shifts"],
                "Shifts per Day": SHIFTS_PER_DAY,
                "Days to Complete": item["days"],
                "Status": "OK",
            }
        )

    summary_rows = [
        {"Metric": "Input rows", "Value": len(input_rows)},
        {"Metric": "Matched rows", "Value": sum(1 for row in match_rows if row["Selected Item Code"])},
        {"Metric": "Descriptions with multiple item-code matches", "Value": sum(1 for row in match_rows if row["Status"].startswith("Multiple"))},
        {"Metric": "Descriptions missing from BOM", "Value": sum(1 for row in match_rows if not row["Selected Item Code"])},
        {"Metric": "Machine-plan item rows after grouping", "Value": len(plan_rows)},
        {"Metric": "Production date", "Value": PRODUCTION_DATE},
        {"Metric": "Shifts per day", "Value": SHIFTS_PER_DAY},
    ]

    return plan_rows, match_rows, summary_rows


def add_sheet(workbook: Workbook, title: str, rows: list[dict]) -> None:
    sheet = workbook.create_sheet(title)
    if not rows:
        return
    headers = list(rows[0].keys())
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])

    header_fill = PatternFill("solid", fgColor="0F6B5E")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "Machine Order": 16,
        "Machine": 24,
        "Process": 24,
        "Item Code": 18,
        "Component": 44,
        "Production Qty": 16,
        "Start Date": 14,
        "End Date": 14,
        "Required Hrs": 14,
        "Status": 36,
        "Item Description": 50,
        "Alternate Item Codes": 44,
        "Metric": 42,
        "Value": 24,
    }
    for idx, header in enumerate(headers, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=idx).column_letter].width = widths.get(header, 18)

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def write_workbook() -> Path:
    plan_rows, match_rows, summary_rows = build_rows()
    workbook = Workbook()
    workbook.remove(workbook.active)
    add_sheet(workbook, "Machine Wise Plan", plan_rows)
    add_sheet(workbook, "Match Review", match_rows)
    add_sheet(workbook, "Input Summary", summary_rows)

    EXPORT_DIR.mkdir(exist_ok=True)
    export_path = EXPORT_DIR / OUTPUT_NAME
    workbook.save(export_path)

    downloads_path = DOWNLOADS_DIR / OUTPUT_NAME
    try:
        workbook.save(downloads_path)
        return downloads_path
    except Exception:
        return export_path


if __name__ == "__main__":
    path = write_workbook()
    print(path)
