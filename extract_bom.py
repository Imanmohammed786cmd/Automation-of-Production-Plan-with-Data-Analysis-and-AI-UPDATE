from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


SOURCE = Path("/Users/imanmohammed/Downloads/Production_Plan_Machine_Automation.xlsx")
OUT = Path("data.js")
TEMPLATE = Path("production_qty_template.csv")


def clean(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return " ".join(value.replace("\xa0", " ").split())
    return value


def main() -> None:
    wb = load_workbook(SOURCE, data_only=True, read_only=True)
    ws = wb["BOM Reference"]

    records = []
    item_seen = {}

    for row in ws.iter_rows(min_row=3, values_only=True):
        item_code, component, material_code, material, part_wt, process, machine, cycle_time, hour_qty = row[:9]
        item_code = str(clean(item_code))
        if not item_code:
            continue

        record = {
            "i": item_code,
            "c": clean(component),
            "mc": str(clean(material_code)),
            "m": clean(material),
            "w": float(part_wt or 0),
            "p": clean(process),
            "ma": clean(machine),
            "ct": float(cycle_time or 0),
            "hq": float(hour_qty or 0),
        }
        records.append(record)
        item_seen.setdefault(item_code, {"code": item_code, "component": record["c"]})

    payload = {
        "source": SOURCE.name,
        "recordCount": len(records),
        "itemCount": len(item_seen),
        "records": records,
        "items": sorted(item_seen.values(), key=lambda x: x["code"]),
    }

    OUT.write_text(
        "window.BOM_DATA = " + json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + ";\n",
        encoding="utf-8",
    )

    sample_codes = list(sorted(item_seen))[:8]
    lines = ["Item Code,Production Qty"]
    for code in sample_codes:
        lines.append(f"{code},1000")
    TEMPLATE.write_text("\n".join(lines) + "\n", encoding="utf-8")

    print(f"Wrote {OUT} with {len(records)} BOM rows and {len(item_seen)} items.")
    print(f"Wrote {TEMPLATE}.")


if __name__ == "__main__":
    main()
